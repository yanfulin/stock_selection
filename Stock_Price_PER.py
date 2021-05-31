import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from datetime import date




def GetPER(stock_id):
    PER_file = Path.cwd() / stock_id / "PER_PBR.html"
    df_PER = pd.read_html(PER_file)[0]
    #print(df_PER.head())
    df_PER["最高PER"]=df_PER["最高PER"].apply(lambda x: x.replace("-", "0"))
    df_PER["最低PER"] = df_PER["最低PER"].apply(lambda x: x.replace("-", "0"))
    df_PER["平均PER"] = df_PER["平均PER"].apply(lambda x: x.replace("-", "0"))
    max_PER=float(df_PER.loc[1,"最高PER"])
    min_PER=float(df_PER.loc[1,"最低PER"])
    avg_PER=float(df_PER.loc[1,"平均PER"])
    return max_PER, min_PER, avg_PER

def GetEPS(stock_id):
    max_PER, min_PER, avg_PER = GetPER(stock_id)
    print("Get the EPS of", stock_id, max_PER, min_PER, avg_PER)
    EPS_file = Path.cwd() / stock_id / "forecast" / "forecast_actual_merged.csv"
    EPS_yearly_file = Path.cwd() / stock_id / "forecast" / "EPS_yearly.csv"
    df_EPS = pd.read_csv(EPS_file)
    EPS_yearly = df_EPS.groupby("Year")["EPS_forecast"].sum()
    EPS_yearly = pd.DataFrame({'Year': EPS_yearly.index, 'EPS': EPS_yearly.values})

    EPS_yearly["max_stock_price"]= (max_PER * EPS_yearly["EPS"])
    EPS_yearly["avg_stock_price"] = (avg_PER * EPS_yearly["EPS"])
    EPS_yearly["min_stock_price"] = (min_PER * EPS_yearly["EPS"])
    EPS_yearly.to_csv(EPS_yearly_file)
    this_year=date.today().year
    three_year_EPS=EPS_yearly.loc[(EPS_yearly["Year"]>=this_year-1),"EPS"].values[0:3]

    #print(three_year_EPS)
    return three_year_EPS


def GetEPS_per_quauter(stock_id):
    EPS_quarterly_file = Path.cwd() / stock_id / "EPS_per_quarter.csv"
    df_EPS = pd.read_csv(EPS_quarterly_file)

    this_year=date.today().year
    two_year_EPS=df_EPS.loc[(df_EPS["Year"]>=this_year-1),"每股稅後盈餘(元)"]
    size = len(two_year_EPS)
    print (two_year_EPS)
    for n in range(size):
        #print(24+n, size-n-1, two_year_EPS[n])
        print(24+n, size-n-1, two_year_EPS[size-n-1])

    #print(three_year_EPS)
    return two_year_EPS, size


def export_to_excel(stock_id):
    Excel_file = Path.cwd() / "stocks.xlsx"
    wb = load_workbook(Excel_file)
    sheet = wb['evaluation']
    # data = sheet.values
    # data =list(data)
    # stock_list2 = [r[0] for r in data]
    # print(stock_list2)



def export_to_excel2():
    Excel_file = Path.cwd() / "stocks.xlsx"
    wb = load_workbook(Excel_file)
    sheet = wb['evaluation']
    #print (sheet.values)
    colnames = ["Stock ID","Chinese Name","Name","Current Price","52W High","52W Low","Beta","Market Cap","PBR","PER",
                "Target Price","Rick's PER","Comment","短中長投資屬性",
                "Dividend","Yield %","Previous Year EPS","This Year EPS","Next Year EPS", "Max PER","Min PER","Avg PER",
                "Price @ High PER","Price @ Avg PER","Price @ Min PER",
                "2010 Q1 EPS", "2010 Q2 EPS", "2010 Q3 EPS", "2010 Q4 EPS",
                "2011 Q1 EPS","2011 Q2 EPS","2011 Q3 EPS","2011 Q4 EPS",
                "2012 Q1 EPS","2012 Q2 EPS","2012 Q3 EPS","2012 Q4 EPS",
                "2010 Total EPS","2011 Total EPS","2012 Total EPS",
                "過去五年平均配息％","股利", "Yield %", "PER High","PER Low","Price @High PE", "Price @Low PE",
                "Price@Low PE- Market Price", "Price @6% Yield"]
    col_indices = {n: cell.value for n, cell in enumerate(sheet['1']) if cell.value in colnames}
    print (col_indices)
    stock_list=[]
    for row in sheet["A"]:
            if row.value != None:
                #print ("row===>", row.value)
                stock_list.append(row.value)
    print(stock_list)
    # stock list: some are in text. some are int. Needs to convert them to right format
    # Get the corresponding EPS and PER and write them into the right position
    for index, stock in enumerate(stock_list, start=1):
        #print(index, stock)
        if index ==1:
            pass
        else:
            stock = str(stock)
            #print(stock)
            EPS = GetEPS(stock)
            PER_max, PER_avg, PER_min = GetPER(stock)
            EPS_quarterly, size=GetEPS_per_quauter(stock)
            sheet.cell(row=index, column=18).value = EPS_quarterly[-4:].sum()
            sheet.cell(row=index, column=19).value = EPS[1]
            sheet.cell(row=index, column=20).value = EPS[2]
            sheet.cell(row=index, column=21).value = PER_max
            sheet.cell(row=index, column=22).value = PER_avg
            sheet.cell(row=index, column=23).value = PER_min

            #Write quarterly EPS. From 2020 Q1 to now (2011)
            for n in range(size):
                sheet.cell(row=index, column=(27+n)).value = EPS_quarterly[size-n-1]

            #sheet.cell(row=index, column=24).value = EPS_quarterly.loc[4]
            
            wb.save(Excel_file)





def main():
    fin = open('StockCode', 'r+')
    StockCodeList = [str(i) for i in fin.read().splitlines()]
    fin.close()
    # print(StockCodeList)

    # for ID in StockCodeList:
    #     print("ID=", ID)
    #     max_PER, min_PER, avg_PER = GetPER(ID)
    #     EPS = GetEPS(ID)
    #     #print (ID, max_PER, min_PER, avg_PER)
    #     print ("2020 EPS=" , EPS)
    export_to_excel2()



if __name__ == "__main__":
    main()